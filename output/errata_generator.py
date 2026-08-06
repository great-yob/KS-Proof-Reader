"""
정오표(正誤表) Excel 생성 모듈
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
교정 적용 결과를 출판사 실무 수준의 Excel 정오표로 출력.

시트 구성 — **적용 결과별로 나눈다**(2026-08-05 개편):
  ✔ 적용      — 문서에 실제로 반영된 자리        ← 결과물 '교정본 + 정오표'
  ✎ 교정      — 사람이 직접 반영해야 할 자리      ← 결과물 '정오표만'
  — 거절      — 사용자가 거절한 제안 (기록용)
  🔍 확인 필요 — 반영 실패 / 사전 검수 플래그 (사람이 직접 손봐야 하는 것)
비어 있는 시트는 만들지 않는다 — 그래서 '적용'과 '교정'은 결과물 옵션에 따라 저절로
한쪽만 남는다. 전부 비면 '적용' 시트만 빈 채로 남는다.

행 단위 = **본문 등장 1곳**. 같은 교정이 12곳에서 반복되면 12행이 되고 각 행이
자기 쪽 번호를 갖는다(반복 중 일부만 수락한 경우도 행마다 정확히 갈린다).

컬럼: 쪽 | 수정 전 | 수정 후 | 교정 이유 | 교정 유형

⚠ '쪽'은 한/글의 **prnpageno**다 — Ctrl+G(쪽 찾아가기)에 그대로 넣을 수 있는 번호이며,
  구역마다 쪽 번호를 새로 시작하는 문서에서는 전역 단조도 유일도 아니다. 정렬 키로 쓰지
  말 것(수집 근거는 core/hwp_bridge_worker._page_no 주석). 비어 있으면 '—' — 브리지가
  못 찾았거나 예산이 소진된 경우로, 표시만 빠질 뿐 행은 그대로 남는다.

색상 범례 (행 배경) — 교정의 **출처**를 나타낸다(유형이 아니다):
  노란색   ← 사전검증 기본
  연두색   ← AI 오탈자 보완
  연보라   ← AI 윤문
  주황색   ← 사전 미등재 주의 항목
  연분홍   ← HWP 매칭 실패 (미적용)

설치:
  pip install openpyxl
"""

import math
import os
import unicodedata
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side,
)

# ── 색상 팔레트 (ARGB) ─────────────────────────────────
_C = {
    # 헤더 배경
    "header_bg":     "FF1F3864",   # 네이비
    "header_fg":     "FFFFFFFF",   # 흰색

    # 타이틀 배경
    "title_bg":      "FF2E4D7B",   # 짙은 파랑
    "title_fg":      "FFFFFFFF",

    # 교정 유형별 행 배경
    "dict":          "FFFFF9C4",   # 연노랑   (사전검증 기본)
    "ai_typo":       "FFE8F5E9",   # 연초록   (AI 오탈자)
    "ai_polish":     "FFEDE7F6",   # 연보라   (AI 윤문)
    "unverified":    "FFFFF3E0",   # 연주황   (사전 미등재 주의)
    "dict_flag":     "FFFFF8E1",   # 연앰버   (사전 검수 — 미등재어 점검)
    "spacing":       "FFE0F2F1",   # 연청록   (띄어쓰기 검수 제안 — 자동수정 아님)
    "punct":         "FFF3E5F5",   # 연자주   (문장부호 검수 제안 — 괄호 짝 등)
    "fail":          "FFFFEBEE",   # 연분홍   (매칭 실패)
    "rejected":      "FFF2F2F2",   # 연회색   (사용자 거절 — 적용 대상 아님)

    # 강조 텍스트
    "ok_fg":         "FF1B5E20",   # 진초록
    "fail_fg":       "FFB71C1C",   # 진빨강
    "warn_fg":       "FFE65100",   # 진주황
    "rej_fg":        "FF9E9E9E",   # 회색     (거절)
    "muted_fg":      "FF888888",   # 옅은 회색(쪽 미확인 등)

    # 테두리
    "border":        "FFB0BEC5",
}

# ── 소스 레이블 ────────────────────────────────────────
_SOURCE_LABEL = {
    "dict":      "사전검증",
    "ai_typo":   "AI 오탈자",
    "ai_polish": "AI 윤문",
    "dict_flag": "사전 검수",
    "spacing":   "띄어쓰기",
    "punct":     "문장부호",
}

# HWP 미등재어 경고 색 (BGR 0x0055FF → 구분 목적)
HL_UNVERIFIED = 0x0055FF

# ── 시트 정의 ──────────────────────────────────────────
#   (시트명, 이 시트가 담는 outcome 집합, 부제)
#   ⚠ 순서가 곧 탭 순서다. '적용'이 첫 시트여야 한다 — 정오표를 여는 사람이 가장 먼저
#     볼 것은 '무엇이 바뀌었나'이기 때문.
#   ⚠ `적용`과 `교정`은 **결과물 옵션에 따라 한쪽만 존재한다** — 교정본 모드는 applied만,
#     '정오표만' 모드는 todo만 나온다. 빈 시트를 만들지 않으므로 모드 인자 없이 이 표
#     하나로 두 모드가 갈린다. ⚠ '적용'은 '이미 문서에 반영됨'을 뜻하므로 정오표만
#     모드에서 그 이름을 쓰면 거짓말이 된다 — 그래서 이름을 나눈 것이다.
_SHEETS = (
    ("적용",      ("applied",),          "문서에 실제로 반영된 자리"),
    ("교정",      ("todo",),             "한글 파일은 수정하지 않았습니다 — 아래 쪽을 찾아 직접 반영하세요"),
    ("거절",      ("rejected",),         "검토 단계에서 거절한 제안 (문서 무변경)"),
    ("확인 필요", ("failed", "review"),  "자동 반영되지 않아 사람이 직접 확인해야 하는 항목"),
)

_HEADERS = ["쪽", "수정 전", "수정 후", "교정 이유", "교정 유형"]
_COL_WIDTHS = {"A": 8, "B": 28, "C": 28, "D": 70, "E": 14}


def generate_errata(
    rows: list,          # 등장 1곳 = 1행짜리 결과 리스트
    hwp_path: str,       # 원본 HWP 파일 경로 (출력 파일명 생성용)
    options: dict,       # 교정 옵션 {"used_ai", "mode", "used_dict"}
    output_path: str = None,  # None이면 HWP 파일과 같은 폴더에 자동 저장
) -> str:
    """
    정오표 Excel 파일 생성.

    Args:
        rows:        [{"page": int|None, "original", "corrected", "reason",
                       "category", "source", "color",
                       "outcome": "applied"|"rejected"|"failed"|"review",
                       "note"}, ...]
                     — ui/workers/apply_worker._build_errata_rows가 만든다.
        hwp_path:    원본 HWP 경로
        options:     {"used_ai": bool, "mode": "typo"|"polish", "used_dict": bool}
        output_path: 저장 경로. None이면 자동 결정.

    Returns:
        생성된 xlsx 파일 경로
    """
    if output_path is None:
        base, _ = os.path.splitext(hwp_path)
        output_path = base + "_정오표.xlsx"

    rows = list(rows or [])
    wb = Workbook()
    wb.remove(wb.active)  # 기본 시트 제거

    made = 0
    for name, outcomes, subtitle in _SHEETS:
        part = [r for r in rows if r.get("outcome") in outcomes]
        # ⚠ 비면 시트를 만들지 않는다 — 빈 '확인 필요' 탭은 '확인할 게 있나?'라는
        #   잘못된 신호를 준다. 단, 전부 비면 최소한 '적용' 한 장은 남긴다(아래).
        if not part:
            continue
        _build_sheet(wb, name, subtitle, part, rows, hwp_path, options)
        made += 1
    if made == 0:
        _build_sheet(wb, "적용", _SHEETS[0][2], [], rows, hwp_path, options)

    wb.save(output_path)
    return output_path


# ══════════════════════════════════════════════════════
# ▌시트 작성
# ══════════════════════════════════════════════════════

def _build_sheet(wb: Workbook, title: str, subtitle: str,
                 part: list, all_rows: list, hwp_path: str, options: dict):
    ws = wb.create_sheet(title)

    for col, width in _COL_WIDTHS.items():
        ws.column_dimensions[col].width = width

    _write_title_block(ws, title, subtitle, part, all_rows, hwp_path, options)

    # ── 헤더 행 (행 4) ────────────────────────────────
    for col_idx, h in enumerate(_HEADERS, start=1):
        cell = ws.cell(row=4, column=col_idx, value=h)
        cell.font      = Font(name="맑은 고딕", bold=True, color=_C["header_fg"], size=10)
        cell.fill      = PatternFill("solid", fgColor=_C["header_bg"])
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = _thin_border()
    ws.row_dimensions[4].height = 24
    ws.freeze_panes = "A5"
    ws.auto_filter.ref = f"A4:E{4 + len(part)}"

    # ── 데이터 행 (행 5~) ─────────────────────────────
    for row_offset, item in enumerate(part):
        row = 5 + row_offset
        source     = item.get("source", "")
        outcome    = item.get("outcome", "")
        is_unverif = (item.get("color", 0) == HL_UNVERIFIED)
        is_review  = (outcome == "review")

        if outcome == "rejected":
            row_bg = _C["rejected"]
        elif outcome == "failed":
            row_bg = _C["fail"]
        elif outcome == "todo":
            # 아직 문서에 없는 지시 — 출처 색을 그대로 살린다(무엇이 잡아낸 교정인지가
            #   손으로 반영할 때도 판단 근거가 된다).
            row_bg = _C["unverified"] if is_unverif else _C.get(source, "FFFFFFFF")
        elif is_review:
            row_bg = _C["dict_flag"]
        elif is_unverif:
            row_bg = _C["unverified"]
        else:
            row_bg = _C.get(source, "FFFFFFFF")

        fill   = PatternFill("solid", fgColor=row_bg)
        border = _thin_border()

        # ① 쪽 — 한/글 쪽 번호. 못 구했으면 '—'(행은 그대로 남긴다).
        page = item.get("page")
        cell = _wcell(ws, row, 1, page if isinstance(page, int) else "—",
                      fill, border,
                      align=Alignment(horizontal="center", vertical="top"))
        if not isinstance(page, int):
            cell.font = Font(name="맑은 고딕", size=9, color=_C["muted_fg"])
        else:
            cell.font = Font(name="맑은 고딕", size=10, bold=True)

        # ② 수정 전
        _wcell(ws, row, 2, item.get("original", ""), fill, border,
               align=Alignment(horizontal="left", vertical="top", wrap_text=True))

        # ③ 수정 후 (치환 없는 검수 플래그는 안내 문구)
        corrected = item.get("corrected", "")
        if is_review and corrected == item.get("original", ""):
            corrected = "(검수 필요 — 표제어 확인)"
        _wcell(ws, row, 3, corrected, fill, border,
               align=Alignment(horizontal="left", vertical="top", wrap_text=True))

        # ④ 교정 이유 — 확인이 필요한 행은 그 사유를 뒤에 잇는다(◆ 구분).
        #   ⚠ 컬럼을 늘리지 않고 여기에 싣는 이유: 다섯 칸 구성은 세 시트가 동일해야
        #     정오표를 통째로 복사·병합해 쓸 수 있다.
        reason = item.get("reason", "") or ""
        note   = item.get("note", "") or ""
        if note and outcome in ("failed", "applied"):
            reason = f"{reason} ◆ {note}" if reason else note
        _wcell(ws, row, 4, reason, fill, border,
               align=Alignment(horizontal="left", vertical="top", wrap_text=True))

        # ⑤ 교정 유형 — 무엇을 고쳤나(맞춤법·띄어쓰기·규범표기…). 비면 출처 라벨.
        _wcell(ws, row, 5, _type_label(item), fill, border,
               align=Alignment(horizontal="center", vertical="top"))

        ws.row_dimensions[row].height = _row_height(
            item.get("original", ""), corrected, reason)

    # ── 범례 (데이터 아래 2행 공백 후) ─────────────────
    legend_row = 5 + len(part) + 2
    ws.cell(row=legend_row, column=1, value="[행 색상 = 교정 출처]").font = \
        Font(name="맑은 고딕", bold=True, size=9, color="FF555555")

    legends = [
        (_C["dict"],       "사전검증 기본"),
        (_C["ai_typo"],    "AI 오탈자 보완"),
        (_C["ai_polish"],  "AI 윤문"),
        (_C["unverified"], "⚠ 사전 미등재 주의 — 사람 검토 필요"),
        (_C["dict_flag"],  "🔍 사전 검수 — 치환 없이 표제어만 확인 요청"),
        (_C["fail"],       "✖ 반영 실패 — 수락했으나 문서에 반영되지 않음 (수동 확인 필요)"),
        (_C["rejected"],   "— 사용자 거절 — 적용 대상 아님"),
    ]
    for i, (color, label) in enumerate(legends):
        r = legend_row + 1 + i
        cell_color = ws.cell(row=r, column=1)
        cell_color.fill   = PatternFill("solid", fgColor=color)
        cell_color.border = _thin_border()
        cell_label = ws.cell(row=r, column=2, value=label)
        cell_label.font      = Font(name="맑은 고딕", size=8, color="FF333333")
        cell_label.alignment = Alignment(vertical="center")
        ws.row_dimensions[r].height = 16

    note_row = legend_row + 1 + len(legends) + 1
    nc = ws.cell(row=note_row, column=1,
                 value="※ '쪽'은 한/글 쪽 번호입니다 — 한/글에서 Alt+G(쪽 찾아가기)에 "
                       "그대로 입력하면 해당 위치로 이동합니다. "
                       "'—'는 쪽 번호를 확인하지 못한 항목입니다.")
    nc.font      = Font(name="맑은 고딕", size=8, color="FF666666")
    nc.alignment = Alignment(vertical="center")
    ws.merge_cells(f"A{note_row}:E{note_row}")


def _write_title_block(ws, title: str, subtitle: str, part: list, all_rows: list,
                       hwp_path: str, options: dict):
    """타이틀 블록 (행 1~3) 작성"""
    doc_name = os.path.basename(hwp_path)
    now_str  = datetime.now().strftime("%Y년 %m월 %d일  %H:%M")
    mode_str = "오탈자·띄어쓰기" if options.get("mode") == "typo" else "전체 윤문"
    stages   = []
    if options.get("deep_screening"): stages.append("표준국어대사전 1차 심층 스크리닝")
    if options.get("used_ai"):        stages.append(f"Gemini AI ({mode_str})")
    if options.get("used_dict"):      stages.append("표준국어대사전 재검증·일관성 가드")

    # 행 1: 제목 — 시트가 담는 결과를 제목에 박는다(탭만 보고는 구분이 약하다).
    ws.merge_cells("A1:E1")
    cell = ws["A1"]
    cell.value     = f"정  오  표  (正誤表)   ·   {title}  {len(part)}항목"
    cell.font      = Font(name="맑은 고딕", bold=True, size=16, color=_C["title_fg"])
    cell.fill      = PatternFill("solid", fgColor=_C["title_bg"])
    cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 36

    # 행 2: 문서 정보 + 다른 시트로 간 항목 수(합계가 어디로 갔는지 한 줄로 보이게).
    tally = []
    for name, outcomes, _sub in _SHEETS:
        n = sum(1 for r in all_rows if r.get("outcome") in outcomes)
        if n:
            tally.append(f"{name} {n}항목")
    ws.merge_cells("A2:E2")
    cell = ws["A2"]
    cell.value = (
        f"  {subtitle}      대상 파일: {doc_name}      교정 일시: {now_str}\n"
        f"  전체: {' · '.join(tally) if tally else '0항목'}"
        + (f"      교정 단계: {' → '.join(stages)}" if stages else "")
    )
    cell.font      = Font(name="맑은 고딕", size=9, color="FF333333")
    cell.fill      = PatternFill("solid", fgColor="FFE8EDF5")
    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.row_dimensions[2].height = 32

    # 행 3: 빈 구분행
    ws.merge_cells("A3:E3")
    ws["A3"].fill = PatternFill("solid", fgColor="FFFFFFFF")
    ws.row_dimensions[3].height = 6


# ══════════════════════════════════════════════════════
# ▌공통 헬퍼
# ══════════════════════════════════════════════════════

# ── 행 높이 ────────────────────────────────────────────
#   ⚠ 예전엔 '수정 전' 글자 수만 보고 높이를 정했다. 그런데 실제로 줄을 잡아먹는 칸은
#     **교정 이유**(폭 70)다 — 대부분 2줄 이상이라 글씨가 잘려 보였다(사용자 보고
#     2026-08-06). 이제 세 텍스트 칸을 각자의 열 폭으로 접어 본 줄 수의 최댓값을 쓴다.
_LINE_PT   = 15.0   # 9pt 맑은 고딕 한 줄 + 여유
_MIN_LINES = 2      # 사용자 지정 — 교정 이유가 통상 2줄이라 최소 2줄로 맞춘다
_MAX_LINES = 8      # 유난히 긴 사유가 한 행을 통째로 차지하지 않도록


def _cell_units(s: str) -> int:
    """엑셀 열 너비 단위('0' 글자 폭)로 잰 문자열 길이 — 한글·한자는 2를 차지한다."""
    return sum(2 if unicodedata.east_asian_width(ch) in ("W", "F") else 1
               for ch in s)


def _wrapped_lines(s, col_width: float) -> int:
    """`wrap_text`가 걸린 칸에서 차지할 줄 수(개행 문자도 반영)."""
    if not s:
        return 1
    usable = max(1.0, col_width - 1)          # 좌우 여백 한 글자분
    return sum(max(1, math.ceil(_cell_units(p) / usable))
               for p in str(s).split("\n"))


def _row_height(original, corrected, reason) -> float:
    lines = max(
        _MIN_LINES,
        _wrapped_lines(original,  _COL_WIDTHS["B"]),
        _wrapped_lines(corrected, _COL_WIDTHS["C"]),
        _wrapped_lines(reason,    _COL_WIDTHS["D"]),
    )
    return min(lines, _MAX_LINES) * _LINE_PT


def _wcell(ws, row, col, value, fill, border, align=None, font=None):
    cell = ws.cell(row=row, column=col, value=value)
    cell.fill   = fill
    cell.border = border
    if align:
        cell.alignment = align
    cell.font = font or Font(name="맑은 고딕", size=9)
    return cell


def _thin_border():
    side = Side(style="thin", color=_C["border"])
    return Border(left=side, right=side, top=side, bottom=side)


def _type_label(item: dict) -> str:
    """교정 유형 — `Correction.category`(맞춤법·띄어쓰기·규범표기·문장부호…)를 쓴다.

    ⚠ 예전 정오표는 '교정 유형'과 '출처'를 두 칸으로 두고 **둘 다 source 라벨**을
      넣어 사실상 같은 값을 두 번 적었다. 유형은 category, 출처는 행 배경색이 맡는다.
    """
    base = (item.get("category") or "").strip()
    if not base:
        base = _SOURCE_LABEL.get(item.get("source", ""), item.get("source", ""))
    return f"⚠ {base}" if item.get("color", 0) == HL_UNVERIFIED else base
